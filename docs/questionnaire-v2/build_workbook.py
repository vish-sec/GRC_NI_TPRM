#!/usr/bin/env python3
"""
Builds NI-TPRM-Controls-v2.xlsx — a detailed bank/FI third-party-risk control
questionnaire mapped to MAS / RBI / SEBI frameworks.

Authoring note: clause references were compiled from web research (June 2026) and
the existing in-app source clause library. They MUST be verified against official
PDFs before use as a system of record. See SOURCES.md and the Readme sheet.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter, OrderedDict

# ---------------------------------------------------------------------------
# Column schema for the Controls sheet
# ---------------------------------------------------------------------------
HEADERS = [
    "Control ID", "Domain", "Sub-domain", "Control Question (asked of vendor)",
    "Intent / Why", "Evidence / RFI expected", "Applicability guidance",
    "Inherent-risk weight", "MAS mapping", "RBI mapping", "SEBI mapping",
    "Recommended cert coverage", "Source reference(s)",
]

# Tuple field order matches HEADERS (minus Control ID which is auto-generated per domain).
# (domain, subdomain, question, intent, evidence, applicability, weight, mas, rbi, sebi, cert, source)
C = []  # control list


def add(domain, sub, q, intent, ev, appl, weight, mas, rbi, sebi, cert, src):
    C.append((domain, sub, q, intent, ev, appl, weight, mas, rbi, sebi, cert, src))


# ===========================================================================
# 1. GOVERNANCE & OUTSOURCING OVERSIGHT
# ===========================================================================
D = "Governance & Outsourcing"
add(D, "Board oversight",
    "Does your board / senior management have documented oversight and ultimate accountability for the security and continuity of the service you provide to us, and can you name the accountable executive?",
    "Confirms governance does not get diluted by outsourcing; the regulated entity remains accountable but expects the provider to have its own board-level ownership.",
    "Org chart, board/management committee charter, RACI naming the accountable executive, last 2 meeting minutes referencing client/security risk.",
    "All vendors handling our systems, data, or customer-facing process.", "High",
    "MAS Guidelines on Outsourcing S5 (board/senior mgmt responsibility)",
    "RBI Outsourcing of FS 2006 5.2; RBI IT Governance 2023 S10",
    "CSCRF GV.SC (RE solely accountable)",
    "ISO 27001 (A.5)", "MAS-OB-GOV; RBI-ITG-S10; SEBI-GVSC-ACCOUNTABLE")
add(D, "Outsourcing register",
    "Will you provide, and keep current, the information we need to maintain our outsourcing register (legal entity, service description, processing locations, sub-contractors, materiality inputs)?",
    "We must maintain a register of outsourced relevant services and submit it to the regulator; we depend on you for accurate inventory inputs.",
    "Completed vendor profile / register data sheet; list of processing locations and sub-processors; notice process for changes.",
    "All vendors; expanded fields for material arrangements.", "High",
    "MAS Notice 658 (register; submission within 15 business days of 30 Jun / 31 Dec)",
    "RBI IT Outsourcing 2023 (outsourcing inventory/materiality)", "CSCRF GV.SC (third-party inventory)",
    "none", "MAS-N658-REGISTER; RBI-ITO-MATERIALITY")
add(D, "Materiality / criticality",
    "Do you understand which of our services you support are material/critical, and do you apply enhanced controls (resilience, audit, reporting) commensurate with that classification?",
    "Material Outsourced Relevant Services (MOORS) and RBI 'material' IT outsourcing attract elevated requirements; the vendor must support them.",
    "Service criticality acknowledgement; control matrix showing enhanced controls for material services.",
    "All; drives applicability of High-weight controls.", "High",
    "MAS Notice 658 (MOORS materiality criteria)", "RBI IT Outsourcing 2023 (materiality of IT outsourcing)",
    "CSCRF GV.SC (assessment proportionate to criticality)", "none",
    "MAS-N658-MOORS; RBI-ITO-MATERIALITY; SEBI-GVSC-RISKASSESS")
add(D, "Outsourcing of regulated/core functions",
    "Do you, or any sub-contractor, perform any core management, compliance, or decision-making function on our behalf, or take decisions that bind us to customers or the regulator?",
    "Core management/compliance/decision-making cannot be outsourced; identifies prohibited delegation early.",
    "Description of decision rights; confirmation no regulated/core function is performed without our explicit authority.",
    "Vendors embedded in lending, KYC, compliance, or customer decisioning.", "High",
    "MAS Guidelines on Outsourcing (responsibilities not delegable)",
    "RBI Outsourcing of FS 2006 5.2 (core functions cannot be outsourced)", "CSCRF GV.SC (accountability retained)",
    "none", "RBI-OFS-NOOUTSOURCE; SEBI-GVSC-ACCOUNTABLE")
add(D, "Conflicts of interest",
    "Do you provide the same/similar service to our direct competitors, and how do you manage resulting conflicts of interest and information-barrier risks?",
    "Identifies conflict, concentration, and confidentiality-leakage exposure across a shared provider.",
    "Conflict-of-interest policy; description of logical/physical separation between client environments.",
    "Shared-platform / SaaS / BPO vendors.", "Med",
    "MAS Guidelines on Outsourcing (conflicts/concentration)", "RBI IT Governance 2023 S10 (conflicts of interest)",
    "CSCRF GV.SC", "none", "RBI-ITG-S10")
add(D, "Provider financial soundness",
    "Can you evidence your financial soundness and going-concern viability for the contract term (audited financials, funding, insurance)?",
    "Financial failure of the provider is a continuity risk; due diligence must assess financial standing.",
    "Last 2 years audited financial statements; cyber/PI insurance certificates; parent guarantees if applicable.",
    "Material vendors; smaller fintech/start-up LSPs.", "Med",
    "MAS Guidelines on Outsourcing S5.4 (financial soundness)", "RBI IT Outsourcing 2023 (due diligence – financial)",
    "CSCRF GV.SC (due diligence)", "none", "MAS-OB-DUEDIL; RBI-ITO-DD")
add(D, "Use of our name / regulatory standing",
    "Do you hold any licence, authorisation, or empanelment relevant to this service (e.g., RBI PA authorisation, CERT-In empanelment), and is it current?",
    "Some services require the provider itself to be authorised/empanelled; confirms legitimacy.",
    "Copy of licence/authorisation/empanelment with validity dates.",
    "Payment aggregators, auditors, cloud/STQC, CERT-In auditors.", "Med",
    "", "RBI PA/PG Directions (PA authorisation); CERT-In empanelment",
    "CSCRF (CERT-In empanelled auditor; MeitY/STQC CSP)", "none", "RBI-PSD-SAR; SEBI-VAPT; SEBI-CLOUD")

# ===========================================================================
# 2. DUE DILIGENCE & CONTRACTS
# ===========================================================================
D = "Due Diligence & Contracts"
add(D, "Initial due diligence",
    "Will you complete our security due-diligence assessment and supporting evidence prior to onboarding and provide remediation timelines for any gaps?",
    "Due diligence on capability, security and controls is required before engaging a provider.",
    "Completed questionnaire, evidence pack, and a dated remediation plan for open items.",
    "All vendors.", "High",
    "MAS Guidelines on Outsourcing S5.4/5.5 (due diligence)",
    "RBI IT Outsourcing 2023 (due diligence); RBI Digital Lending 2025 (LSP/DLA DD)",
    "CSCRF GV.SC (pre-onboarding risk assessment)", "ISO 27001 / SOC 2",
    "MAS-OB-DUEDIL; RBI-ITO-DD; RBI-DL-DD; SEBI-GVSC-RISKASSESS")
add(D, "Periodic re-assessment",
    "Do you accept periodic re-due-diligence (at least every 24 months for material services, or on significant change) including updated evidence and audit reports?",
    "Provider risk must be re-assessed periodically across the lifecycle, not just at onboarding.",
    "Agreement to recurring re-assessment; cadence in contract; refreshed certifications.",
    "Material vendors at least 24-monthly; others risk-based.", "Med",
    "MAS Guidelines on Outsourcing (periodic re-DD <=24 months)",
    "RBI IT Outsourcing 2023 (periodic DD); RBI Digital Lending 2025 (periodic LSP DD)",
    "CSCRF GV.SC (periodic assessment)", "none", "MAS-OB-REDD; RBI-DL-DD")
add(D, "Binding written agreement",
    "Is there a single, legally enforceable written agreement covering scope, SLAs, confidentiality/security, audit rights, sub-contracting, BCM, liability, termination/exit, and governing law?",
    "Outsourcing must rest on a binding agreement addressing the mandated terms.",
    "Executed master agreement + schedules; clause-mapping matrix to required terms.",
    "All vendors.", "High",
    "MAS Guidelines on Outsourcing S5.6 (agreement terms)",
    "RBI IT Outsourcing 2023 (agreement schedule); RBI Outsourcing of FS 2006 5.5",
    "CSCRF GV.SC (contractual clauses)", "none",
    "MAS-OB-AGREEMENT; RBI-ITO-AGREEMENT; RBI-OFS-CONTRACT; SEBI-GVSC-CONTRACT")
add(D, "Governing law / jurisdiction",
    "Is the agreement governed by the law and courts of the jurisdiction we require (e.g., India for RBI-regulated services), and are our and the regulator's rights enforceable thereunder?",
    "Indian-governing-law and enforceable regulator access are explicit RBI expectations.",
    "Governing-law and jurisdiction clauses; legal opinion for cross-border arrangements.",
    "RBI-regulated services; cross-border vendors.", "Med",
    "MAS Guidelines on Outsourcing S5.8 (cross-border legal risk)",
    "RBI IT Outsourcing 2023 (Indian governing law)", "CSCRF GV.SC", "none",
    "RBI-ITO-AGREEMENT; MAS-OB-CROSSBORDER")
add(D, "Right-to-audit clause",
    "Does the contract grant us, our appointed auditors, and our regulator the right to audit and inspect your records, systems and premises (and those of material sub-contractors)?",
    "Audit/inspection access for the institution, its auditors and the regulator is mandatory.",
    "Audit-rights clause text; confirmation it flows down to sub-contractors.",
    "All material vendors.", "High",
    "MAS Notice 658 (audit/inspection access)", "RBI IT Outsourcing 2023 (audit rights); RBI Outsourcing of FS 5.5.1",
    "CSCRF GV.SC (right to audit)", "none", "MAS-N658-AUDIT; RBI-ITO-AUDIT; SEBI-GVSC-AUDITRIGHTS")
add(D, "Regulator inspection rights",
    "Do you acknowledge and contractually preserve the regulator's right to access and inspect your documents, records and premises relating to our service?",
    "The regulator must retain direct inspection rights even though the function is outsourced.",
    "Regulator-access clause; precedent of facilitating a regulator/auditor visit if any.",
    "Material vendors.", "High",
    "MAS Notice 658 (MAS access)", "RBI IT Outsourcing 2023 (RBI right to inspect); RBI Outsourcing of FS 5.5.1",
    "CSCRF GV.SC", "none", "RBI-ITO-RBIINSPECT; RBI-OFS-RBIINSPECT; MAS-N658-AUDIT")
add(D, "Liability & indemnity",
    "Do the contract's liability, indemnity and breach-remedy provisions adequately cover security incidents, data breaches, and regulatory penalties attributable to you?",
    "Allocates financial responsibility for security failures and regulatory consequences.",
    "Liability/indemnity clauses; breach cost-allocation; insurance backing.",
    "Material vendors.", "Med",
    "MAS Guidelines on Outsourcing S5.6 (liabilities)", "RBI IT Outsourcing 2023; RBI Outsourcing of FS 5.5",
    "CSCRF GV.SC", "none", "MAS-OB-AGREEMENT; RBI-OFS-CONTRACT")
add(D, "SLA & service-level reporting",
    "Are measurable SLAs defined (availability, response, security remediation timelines) with regular reporting and remedies for breach?",
    "Enables ongoing monitoring of performance and control effectiveness.",
    "SLA schedule; sample monthly service/security report; penalty/credit regime.",
    "All vendors providing an operational service.", "Med",
    "MAS Guidelines on Outsourcing (SLAs/monitoring)", "RBI IT Outsourcing 2023 (SLAs); RBI IT Governance 2023 S12",
    "CSCRF GV.SC", "none", "MAS-OB-AGREEMENT; RBI-ITG-S12-SUPPORT")
add(D, "Ongoing monitoring",
    "Do you support our ongoing monitoring through periodic security reports, KPI/KRI dashboards, attestations, and notification of material control changes?",
    "Continuous monitoring of performance, control effectiveness and risk profile is required.",
    "Reporting cadence; sample KRI report; change-notification procedure.",
    "Material vendors.", "Med",
    "MAS Guidelines on Outsourcing (ongoing monitoring)", "RBI IT Outsourcing 2023 (monitoring/control)",
    "CSCRF GV.SC (ongoing monitoring)", "none", "MAS-OB-MONITOR; RBI-ITG-ISAUDIT")
add(D, "Personnel background screening",
    "Do you perform background verification on personnel (and sub-contractor staff) who can access our systems or data, and require confidentiality/NDA agreements?",
    "Insider risk control; provider must vet staff and sub-contractors with access.",
    "Background-check policy; sample screening attestation; signed NDAs.",
    "Vendors with personnel access to our data/systems.", "Med",
    "MAS Guidelines on Outsourcing (personnel controls)",
    "RBI IT Outsourcing 2023 (DD on staff & sub-contractors); RBI NBFC IT Framework 2017",
    "CSCRF GV.SC", "ISO 27001 (A.6)", "RBI-ITO-DD-STAFF; RBI-NBFC-VENDORBG")
add(D, "Security awareness training",
    "Do personnel handling our systems/data receive regular security awareness and role-specific training, and can you evidence completion?",
    "Human-factor control; training must extend to vendor personnel handling bank data.",
    "Training policy/curriculum; completion records; phishing-simulation results.",
    "Vendors with personnel access.", "Low",
    "MAS TRM 2021 (people/awareness)", "RBI CSF 2016 Annex 23.1 (awareness extends to vendors)",
    "CSCRF (awareness)", "ISO 27001 (A.6.3)", "RBI-CSF-AWARENESS")

# ===========================================================================
# 3. DATA SECURITY & RESIDENCY
# ===========================================================================
D = "Data Security & Residency"
add(D, "Data classification & inventory",
    "Do you maintain a classified inventory of our data (including customer/personal/payment data) and apply protection commensurate with classification?",
    "Protection must be driven by classification; you cannot protect data you have not inventoried.",
    "Data inventory/data-flow diagram for our data; classification policy.",
    "All vendors processing our data.", "High",
    "MAS TRM 2021 S12 (classified asset inventory)", "RBI CSF 2016 C8 (data protection)",
    "CSCRF PR.DS", "ISO 27001 (A.5.12)", "MAS-TRM-DATASEC; SEBI-PRDS-ENC")
add(D, "Encryption at rest",
    "Is our data encrypted at rest using strong, industry-standard algorithms (e.g., AES-256), and how is this enforced across storage, backups and removable media?",
    "Protects confidentiality of data at rest including in backups and on lost media.",
    "Encryption standard/config; scope (DB, files, backups); algorithm/key length evidence.",
    "All vendors storing our data.", "High",
    "MAS TRM 2021 S12 (data at rest)", "RBI CSF 2016 C8 (encryption at rest)",
    "CSCRF PR.DS (AES-256-class at rest)", "ISO 27001 / PCI DSS", "MAS-TRM-CRYPTO; RBI-CSF-ACCESS-ENC; SEBI-PRDS-ENC")
add(D, "Encryption in transit",
    "Is all transmission of our data protected with strong TLS (TLS 1.2+), with weak protocols/ciphers disabled and certificate validation enforced?",
    "Protects data moving between us, you, and your sub-processors / APIs.",
    "TLS configuration / SSL Labs-style scan; cipher suite list; cert management process.",
    "All vendors transmitting our data.", "High",
    "MAS TRM 2021 S11/S12 (in transit)", "RBI CSF 2016 C8 (encryption in transit)",
    "CSCRF PR.DS (strong TLS in transit)", "ISO 27001 / PCI DSS", "MAS-TRM-API; SEBI-PRDS-ENC")
add(D, "Data residency – India (RBI)",
    "For RBI-regulated services, is all of our data stored and processed within India, and retrievable in India, including by any sub-contractor?",
    "RBI requires RE data to be stored in India and retrievable in India.",
    "Hosting region attestation; data-flow showing India residency; sub-processor locations.",
    "RBI-regulated services (IT outsourcing, lending, NBFC).", "High",
    "", "RBI IT Outsourcing 2023 (data storage in India); RBI Digital Lending 2025 (India-only storage)",
    "CSCRF PR.DS (data localisation – currently in abeyance, verify status)", "none",
    "RBI-ITO-DATALOC; RBI-DL-DATAMIN")
add(D, "Payment data localisation",
    "For payment-system data, is the entire end-to-end payment data stored only in India, with any foreign-leg processing data deleted and brought back within 24 hours?",
    "RBI's payment-data localisation rule requires full payment data in India; foreign-leg purge within 24h.",
    "Payment data-flow; India-storage attestation; 24-hour purge evidence; SAR.",
    "Payment system providers/intermediaries.", "High",
    "", "RBI Storage of Payment System Data 2018 (entire data in India); RBI PSD FAQ (24-hour return)",
    "", "PCI DSS", "RBI-PSD-LOC; RBI-PSD-24HR")
add(D, "Data minimisation (lending)",
    "Do you collect only need-based data with borrower consent and, as an LSP/DLA, restrict stored data to the minimum basic fields permitted (name, address, contact)?",
    "Digital-lending rules cap LSP-stored data and require need-based, consent-driven collection.",
    "Data-collection matrix vs. minimum-necessary; consent capture; LSP storage scope.",
    "LSPs / DLAs in digital lending.", "High",
    "", "RBI Digital Lending 2025 (data minimisation; LSP minimal data; India storage)",
    "", "none", "RBI-DL-DATAMIN")
add(D, "No biometric storage (lending)",
    "Do you confirm that no borrower biometric data is stored by you or your sub-contractors, except where specifically mandated by statute/regulation?",
    "Digital-lending rules prohibit storage of borrower biometric data.",
    "Written confirmation; data inventory showing absence of biometric storage.",
    "LSPs / DLAs handling biometrics.", "High",
    "", "RBI Digital Lending 2025 (no biometric storage)", "", "none", "RBI-DL-BIOMETRIC")
add(D, "Keys/key-mgmt within India (SEBI)",
    "For SEBI-regulated services, are encryption keys and all key-management operations for our data performed and kept within India?",
    "CSCRF requires keys and key-management within India (status: data-localisation provisions partly in abeyance — verify).",
    "Key-location attestation; HSM/KMS region; key-management runbook.",
    "SEBI-regulated services.", "High",
    "", "", "CSCRF PR.DS (keys within India; Aug 2025 clarifications; localisation in abeyance)", "none",
    "SEBI-PRDS-KEYLOC")
add(D, "Cross-border transfer safeguards",
    "Where any of our data is transferred or processed outside the home jurisdiction, what legally enforceable safeguards and encryption protect it, and is regulator/auditor access preserved?",
    "Cross-border processing must keep comparable protection, encryption and enforceable access rights.",
    "Transfer-impact assessment; SCC/DPA; encryption-in-transit/at-rest for offshore data.",
    "Cross-border arrangements.", "High",
    "MAS Guidelines on Outsourcing S5.8 (cross-border); PDPA S26 (transfer limitation)",
    "RBI IT Outsourcing 2023 (cross-border country risk)", "CSCRF PR.DS / GV.SC", "none",
    "MAS-OB-CROSSBORDER; MAS-XB-ENCRYPTION; MAS-PDPA-XBORDER; RBI-ITO-CROSSBORDER")
add(D, "Data location disclosure",
    "Will you disclose and keep updated all locations (including sub-contractors) where our data is stored, processed or accessed?",
    "Location transparency is needed for residency, country-risk and audit-access assessment.",
    "Current list of processing locations; change-notification process.",
    "All vendors processing our data.", "Med",
    "MAS Outsourcing/TRM (location disclosure)", "RBI IT Outsourcing 2023 (cross-border)",
    "CSCRF GV.SC", "none", "MAS-XB-RESIDENCY; RBI-ITO-CROSSBORDER")
add(D, "Multi-tenant data segregation",
    "In multi-tenant/shared environments, how is our data logically (or physically) segregated from other clients' data to prevent cross-tenant access or leakage?",
    "Prevents confidentiality breach via tenant isolation failure in SaaS/cloud.",
    "Tenant-isolation architecture; access-control model; penetration-test of isolation.",
    "SaaS / multi-tenant cloud vendors.", "High",
    "MAS Guidelines on Outsourcing S5.6 (data segregation)", "RBI IT Outsourcing 2023 (data protection)",
    "CSCRF PR.DS", "ISO 27001 / SOC 2", "MAS-OB-CONFID")
add(D, "Data loss prevention (DLP)",
    "Do you operate DLP controls over our data, and do these extend to your facilities and sub-contractors handling our data?",
    "DLP must extend to vendor facilities storing/processing bank data.",
    "DLP policy/coverage; egress-control evidence; coverage of sub-contractor sites.",
    "Vendors processing sensitive/customer data.", "Med",
    "MAS TRM 2021 S12 (prevent data loss)", "RBI CSF 2016 15.3 (DLP extends to vendor)",
    "CSCRF PR.DS", "ISO 27001", "MAS-TRM-DATASEC; RBI-CSF-DLP")
add(D, "Secure media disposal / data return",
    "How is our data securely destroyed or returned at end of life / contract, with certificates of destruction, including from backups and sub-contractors?",
    "Prevents residual-data exposure after disposal or termination.",
    "Media-sanitisation standard; certificate of destruction template; backup-purge process.",
    "All vendors storing our data.", "Med",
    "MAS TRM 2021 S12 (secure media disposal)", "RBI IT Outsourcing 2023 (exit – data return/destruction)",
    "CSCRF GV.SC (secure exit)", "ISO 27001 (A.7.14/8.10)", "MAS-TRM-DATASEC; RBI-ITO-EXIT")
add(D, "Confidentiality & secrecy",
    "How do you protect the confidentiality of our and our customers' information, restricting access to need-to-know and complying with banking-secrecy obligations?",
    "Confidentiality/secrecy obligations apply where customer information is disclosed to a provider.",
    "Confidentiality policy; access-restriction model; secrecy-compliance attestation.",
    "Vendors with customer-information access.", "High",
    "MAS Guidelines on Outsourcing S5.6 (confidentiality); Banking Act secrecy",
    "RBI Outsourcing of FS 5.6 (breach/leakage)", "CSCRF GV.SC (data protection/confidentiality)",
    "ISO 27001", "MAS-OB-CONFID; MAS-OB-CONSENT; RBI-OFS-BREACH")

# ===========================================================================
# 4. CRYPTOGRAPHY & KEY MANAGEMENT
# ===========================================================================
D = "Cryptography & Key Mgmt"
add(D, "Crypto standards / policy",
    "Do you maintain a cryptographic policy specifying approved algorithms, key lengths and protocols, and prohibiting deprecated cryptography (e.g., MD5, SHA-1, SSLv3/TLS1.0)?",
    "Ensures strong, current cryptography and removal of broken primitives.",
    "Crypto policy; approved-algorithm list; deprecation schedule.",
    "All vendors using cryptography on our data.", "Med",
    "MAS TRM 2021 S12 (strong cryptography)", "RBI CSF 2016 C8", "CSCRF PR.DS", "ISO 27001 (A.8.24)",
    "MAS-TRM-CRYPTO; SEBI-PRDS-ENC")
add(D, "Key lifecycle management",
    "How are cryptographic keys generated, stored, distributed, rotated, revoked and destroyed across their lifecycle, and are they protected in HSMs?",
    "A robust key lifecycle is integral to effective encryption.",
    "Key-management procedure; rotation schedule; HSM/KMS evidence; FIPS 140-2/3 attestation.",
    "Vendors performing encryption/signing on our data.", "High",
    "MAS TRM 2021 S12 (key lifecycle)", "RBI CSF 2016 C8", "CSCRF PR.DS", "ISO 27001 / PCI DSS",
    "MAS-TRM-CRYPTO; SEBI-PRDS-ENC")
add(D, "Customer control of keys (cloud)",
    "For cloud-hosted RE data, do encryption keys and HSMs remain under our control, with the ability to manage and revoke them independently of the cloud provider?",
    "RBI cloud appendix requires keys/HSMs under RE control with independent revocation.",
    "Key-ownership/BYOK or HYOK architecture; revocation runbook.",
    "Cloud-hosted RBI-regulated data.", "High",
    "", "RBI IT Outsourcing 2023 Appendix I Cloud (keys/HSM under RE control)",
    "CSCRF PR.DS (keys in India)", "none", "RBI-ITO-CLOUDKEY; SEBI-PRDS-KEYLOC")
add(D, "Certificate management",
    "How do you manage TLS/code-signing certificates (issuance, inventory, expiry monitoring, revocation) to prevent outages and impersonation?",
    "Expired/mismanaged certificates cause outages and enable interception.",
    "Certificate inventory; expiry-monitoring; CA/issuance policy.",
    "Vendors operating TLS endpoints/APIs for us.", "Low",
    "MAS TRM 2021 S11 (secure connectivity)", "RBI CSF 2016 C8", "CSCRF PR.DS", "ISO 27001",
    "MAS-TRM-API")
add(D, "Secrets management",
    "How are application secrets, API keys and service credentials stored and rotated (e.g., vault/secret manager), and are secrets kept out of source code and config files?",
    "Hard-coded/leaked secrets are a common breach vector across the supply chain.",
    "Secrets-management tooling; rotation policy; secret-scanning in CI.",
    "Vendors integrating with our APIs/systems.", "Med",
    "MAS TRM 2021 S6/S9 (secure development/access)", "RBI IT Governance 2023 S13",
    "CSCRF PR.AA / PR.DS", "ISO 27001 / SOC 2", "MAS-TRM-CRYPTO; MAS-TRM-ACCESS")

# ===========================================================================
# 5. ACCESS CONTROL & MFA
# ===========================================================================
D = "Access Control & MFA"
add(D, "Least-privilege & RBAC",
    "Do you enforce role-based, least-privilege/need-to-know access to systems and data that touch our environment, with documented authorisation?",
    "Limits blast radius of compromised accounts and insider misuse.",
    "Access-control policy; RBAC role matrix; sample access-grant approval.",
    "All vendors with access to our data/systems.", "High",
    "MAS TRM 2021 S9 (least-privilege)", "RBI CSF 2016 C8; RBI IT Outsourcing 2023 Cloud (RBAC)",
    "CSCRF PR.AA (least privilege)", "ISO 27001 (A.5.15) / SOC 2", "MAS-TRM-ACCESS; RBI-ITO-CLOUDRBAC; SEBI-PRAA-MFA")
add(D, "MFA for privileged/remote access",
    "Is multi-factor authentication enforced for all administrative/privileged accounts and for any remote or internet-facing access to systems handling our data?",
    "MFA is mandatory for admin accounts and internet-facing access to customer data.",
    "MFA policy; coverage report for admin/remote access; MFA technology in use.",
    "All vendors with privileged/remote/internet access.", "High",
    "MAS Notice 655 4.6/4.7 (MFA admin & internet-facing customer data)",
    "RBI IT Outsourcing 2023 Cloud (MFA); RBI CSF 2016 C8",
    "CSCRF PR.AA (MFA for critical/admin)", "ISO 27001 / SOC 2", "MAS-655-MFA; RBI-ITO-CLOUDRBAC; SEBI-PRAA-MFA")
add(D, "Privileged access management (PAM)",
    "Do you manage privileged accounts through a PAM solution (vaulting, session recording, just-in-time access, periodic review)?",
    "Privileged accounts are high-value targets requiring strict control and monitoring.",
    "PAM architecture; session-recording sample; JIT/check-out evidence.",
    "Vendors with privileged access to our environment.", "High",
    "MAS TRM 2021 S9 (privileged accounts)", "RBI CSF 2016 C8", "CSCRF PR.AA (PAM)", "ISO 27001 / SOC 2",
    "MAS-TRM-ACCESS; SEBI-PRAA-MFA")
add(D, "Administrative account security",
    "How are administrative accounts on operating systems and databases secured, and do you prohibit using admin accounts for non-administrative activity?",
    "Admin accounts must be secured against unauthorised use and not used for routine work.",
    "Admin-account policy; segregation of admin vs. standard accounts; named-account model.",
    "Vendors operating OS/DB on our behalf.", "Med",
    "MAS Notice 655 4.1 (admin accounts)", "RBI CSF 2016 C5 (remove default creds)",
    "CSCRF PR.AA", "ISO 27001", "MAS-655-ADMIN; RBI-CSF-SECCONFIG")
add(D, "Joiner/mover/leaver & access review",
    "Do you have a JML process and perform periodic access recertification to remove stale/excessive access, especially for accounts touching our data?",
    "Stale access and orphaned accounts are a leading cause of unauthorised access.",
    "JML procedure; last access-recertification report; deprovisioning SLAs.",
    "Vendors with personnel access.", "Med",
    "MAS TRM 2021 S9 (access review)", "RBI CSF 2016 C8", "CSCRF PR.AA", "ISO 27001 / SOC 2",
    "MAS-TRM-ACCESS")
add(D, "Segregation of duties",
    "Are conflicting duties (e.g., development vs. production deployment, request vs. approval) segregated for staff operating our service?",
    "SoD prevents single-person fraud/error from causing undetected harm.",
    "SoD matrix; toxic-combination analysis; compensating controls where SoD not feasible.",
    "Vendors with operational/dev access.", "Med",
    "MAS TRM 2021 S9 (segregation of duties)", "RBI IT Governance 2023 S13", "CSCRF PR.AA",
    "ISO 27001 / SOC 2", "MAS-TRM-ACCESS")
add(D, "Authentication policy",
    "What password/credential policy applies (length, complexity, rotation, lockout, prohibition of shared accounts) for access to systems handling our data?",
    "Weak/shared credentials undermine all other access controls.",
    "Authentication policy; lockout config; shared-account exception register.",
    "All vendors with system access.", "Low",
    "MAS Notice 655 (admin auth)", "RBI CSF 2016 C8", "CSCRF PR.AA", "ISO 27001", "MAS-655-ADMIN")

# ===========================================================================
# 6. NETWORK & MALWARE
# ===========================================================================
D = "Network & Malware"
add(D, "Network perimeter defence",
    "What perimeter controls (firewalls, IPS, WAF, DDoS protection) restrict unauthorised traffic to systems hosting our service, and how are rule-sets reviewed?",
    "Perimeter controls restrict unauthorised access and detect/block malicious traffic.",
    "Network architecture/segmentation diagram; firewall-review cadence; WAF/DDoS evidence.",
    "Internet-facing vendors hosting our service.", "Med",
    "MAS Notice 655 4.4 (perimeter defence)", "RBI CSF 2016 C5/C8", "CSCRF PR (network protection)",
    "ISO 27001 / SOC 2", "MAS-655-PERIMETER")
add(D, "Network segmentation",
    "Is the environment hosting our data segmented from corporate/other-client networks, with controlled, monitored east-west traffic?",
    "Segmentation/micro-segmentation limits lateral movement after a breach.",
    "Micro-segmentation/zoning diagram; inter-zone ruleset; data-flow restrictions.",
    "Vendors hosting our data.", "Med",
    "MAS TRM 2021 S11/S12 (infrastructure security)", "RBI CSF 2016 C5", "CSCRF PR",
    "ISO 27001", "MAS-TRM-DATASEC")
add(D, "Malware protection",
    "Is anti-malware/EDR deployed on all relevant endpoints and servers in the environment serving us, with current signatures/behavioural detection and alerting?",
    "Malware protection is required on every system where available.",
    "EDR/AV coverage report; detection/response config; sample alert handling.",
    "All vendors operating endpoints/servers for us.", "Med",
    "MAS Notice 655 4.5 (malware protection)", "RBI CSF 2016 C5/C7", "CSCRF DE.CM",
    "ISO 27001 / SOC 2", "MAS-655-MALWARE")
add(D, "Secure configuration / hardening",
    "Do you harden systems to documented baselines (e.g., CIS benchmarks), remove default credentials, and verify conformance?",
    "Hardened baselines reduce attack surface; defaults are a known weakness.",
    "Hardening standards; baseline-conformance scan; default-credential removal evidence.",
    "Vendors operating systems for us.", "Med",
    "MAS Notice 655 4.3 (security standards/hardening)", "RBI CSF 2016 C5 (secure configuration)",
    "CSCRF PR", "ISO 27001 (A.8.9)", "MAS-655-STANDARDS; RBI-CSF-SECCONFIG")
add(D, "Secure API/system connectivity",
    "How is connectivity between our systems and yours secured (mutual authentication, encryption, input validation, IP allow-listing, API gateway)?",
    "API/integration points are a direct supply-chain attack path.",
    "API security design; mTLS/auth config; input-validation controls; gateway policy.",
    "Vendors integrating via API to our systems.", "High",
    "MAS TRM 2021 S11 (secure connectivity/API)", "RBI IT Governance 2023 S13", "CSCRF PR",
    "ISO 27001 / SOC 2", "MAS-TRM-API")
add(D, "API rate limiting & abuse protection",
    "Do APIs we consume/expose enforce rate-limiting, throttling and anomaly detection to resist abuse, credential-stuffing and DoS?",
    "Prevents API abuse and resource-exhaustion against shared interfaces.",
    "Rate-limit/throttle config; abuse-detection alerts; quota policy.",
    "API vendors.", "Med",
    "MAS TRM 2021 S11", "RBI IT Governance 2023 S13", "CSCRF PR", "ISO 27001", "MAS-TRM-API")
add(D, "Remote access security",
    "How is remote/administrative access to the environment serving us controlled (VPN/ZTNA, MFA, jump hosts, session logging)?",
    "Remote access is a frequent initial-access vector for breaches.",
    "Remote-access architecture; MFA enforcement; bastion/jump-host logging.",
    "Vendors with remote admin access.", "Med",
    "MAS TRM 2021 S9 (access)", "RBI IT Outsourcing 2023 Cloud (remote access MFA)",
    "CSCRF PR.AA", "ISO 27001 / SOC 2", "MAS-TRM-ACCESS; RBI-ITO-CLOUDRBAC")

# ===========================================================================
# 7. VULNERABILITY & PATCH MANAGEMENT
# ===========================================================================
D = "Vulnerability & Patch"
add(D, "Patch management process",
    "Do you have a risk-based patch-management process that identifies, tests, prioritises and applies security patches within defined timelines (or applies compensating controls)?",
    "Timely patching closes known exploited vulnerabilities; compensating controls cover gaps.",
    "Patch-management policy with SLAs; recent patch-compliance report.",
    "Vendors operating systems/software for us.", "High",
    "MAS Notice 655 4.2; MAS TRM 2021 S10 (patch mgmt)", "RBI CSF 2016 C7; RBI IT Outsourcing 2023 Cloud (patch)",
    "CSCRF PR.MA", "ISO 27001 / SOC 2", "MAS-655-PATCH; MAS-TRM-PATCH; RBI-CSF-PATCHVULN; RBI-ITO-CLOUDPATCH")
add(D, "Patch remediation timelines",
    "What are your remediation SLAs for critical/high vulnerabilities, and can you evidence meeting them (e.g., critical within ~1 week)?",
    "Defined, met SLAs are the measurable proof of effective remediation.",
    "Remediation-SLA matrix; aging report for open critical/high findings.",
    "All vendors with internet-facing/critical systems.", "High",
    "MAS TRM 2021 S14", "RBI IT Governance 2023 (remediation timelines)",
    "CSCRF PR.MA (high/critical within 1 week; others 3 months)", "none", "SEBI-PRMA-PATCH; MAS-TRM-VAPT")
add(D, "Vulnerability scanning",
    "Do you perform regular internal and external vulnerability scanning of the environment serving us, and track findings to closure?",
    "Continuous scanning surfaces new weaknesses between point-in-time pen tests.",
    "Scan cadence; sample scan report (sanitised); remediation tracker.",
    "Vendors operating systems for us.", "Med",
    "MAS TRM 2021 S14 (VA)", "RBI CSF 2016 C7; RBI IT Governance 2023 (semi-annual VA)",
    "CSCRF VAPT", "ISO 27001 / SOC 2", "MAS-TRM-VAPT; RBI-CSF-PATCHVULN")
add(D, "Threat & vulnerability intelligence",
    "Do you consume threat/vulnerability intelligence to prioritise emergency patching for actively exploited vulnerabilities affecting our service?",
    "Risk-based prioritisation requires awareness of active exploitation.",
    "Threat-intel sources; emergency-patch process; example zero-day response.",
    "Material/critical vendors.", "Low",
    "MAS TRM 2021 S13 (threat intelligence)", "RBI IT Governance 2023", "CSCRF DE.CM",
    "none", "MAS-TRM-SOC")
add(D, "Asset & EoL management",
    "Do you maintain an asset inventory for systems serving us and manage end-of-life/unsupported components (no unsupported OS/software in scope)?",
    "Unsupported components cannot be patched and are a persistent risk.",
    "Asset inventory; EoL register; replacement plan for unsupported components.",
    "Vendors operating infrastructure for us.", "Med",
    "MAS TRM 2021 S12 (asset inventory)", "RBI CSF 2016 C5", "CSCRF ID/PR", "ISO 27001 (A.5.9)",
    "MAS-TRM-DATASEC")

# ===========================================================================
# 8. SDLC & SUPPLY CHAIN (SBOM)
# ===========================================================================
D = "SDLC & Supply Chain (SBOM)"
add(D, "Secure SDLC",
    "Do you follow a secure SDLC with security requirements, secure-coding standards, peer review, and security testing (SAST/DAST) before release?",
    "Security must be built in; secure SDLC reduces vulnerabilities in delivered software.",
    "SDLC/secure-coding policy; SAST/DAST evidence; gate criteria.",
    "Vendors developing software/services for us.", "High",
    "MAS TRM 2021 S6 (secure SDLC)", "RBI CSF 2016 6.2 (app security)", "CSCRF (secure development)",
    "ISO 27001 (A.8.25-8.28) / SOC 2", "MAS-TRM-SDLC; RBI-CSF-APPSEC")
add(D, "Source code / security review of software",
    "Is vendor-supplied software subjected to source-code review or equivalent assurance to detect vulnerabilities, malicious code and back-doors before deployment?",
    "Source-code/security review of vendor software detects hidden flaws and tampering.",
    "Code-review/assurance reports; SAST results; third-party assessment summary.",
    "Vendors supplying software/applications.", "High",
    "MAS TRM 2021 S6/S14 (source code review)", "RBI CSF 2016 6.2; RBI IT Governance 2023 S12",
    "CSCRF (secure development)", "ISO 27001", "MAS-TRM-SOURCECODE; RBI-CSF-APPSEC")
add(D, "Vulnerability-free / clean-code certification",
    "Will you certify in writing that delivered applications are free of known vulnerabilities, malware, and covert channels / back-doors at delivery and after major changes?",
    "RBI IT Governance requires vendor written certification of clean, vulnerability-free software.",
    "Signed certification per release; supporting scan evidence.",
    "Vendors supplying critical applications.", "High",
    "MAS TRM 2021 S6", "RBI IT Governance 2023 S12 (vulnerability-free certification)",
    "CSCRF GV.SC", "none", "RBI-ITG-VULNCERT")
add(D, "Software Bill of Materials (SBOM)",
    "Will you provide and maintain an SBOM for software used in our core/critical operations — at purchase and on every change — with the prescribed minimum fields?",
    "SBOM enables rapid identification of vulnerable components across the supply chain.",
    "SBOM (e.g., CycloneDX/SPDX) per release; update-on-change commitment; field completeness.",
    "Vendors supplying software for core/critical ops.", "High",
    "MAS TRM 2021 S6 (software assurance)", "RBI IT Governance 2023 (SBOM-aligned)",
    "CSCRF SBOM (min fields; board-approved exceptions; Aug 2025 clarifications)",
    "none", "SEBI-SBOM")
add(D, "Open-source / dependency management",
    "How do you manage open-source and third-party dependencies (license compliance, known-vulnerability scanning, update cadence) in software delivered to us?",
    "Vulnerable transitive dependencies are a primary supply-chain risk vector.",
    "SCA (software composition analysis) reports; dependency-update policy; license inventory.",
    "Vendors delivering software with OSS components.", "Med",
    "MAS TRM 2021 S6", "RBI IT Governance 2023 (supply-chain)", "CSCRF SBOM / GV.SC",
    "ISO 27001", "SEBI-SBOM; RBI-ITG-S10")
add(D, "Build/CI-CD pipeline security",
    "How is your build/CI-CD pipeline secured against tampering (signed artifacts, integrity verification, restricted pipeline access, provenance)?",
    "Pipeline compromise can inject malicious code into otherwise-trusted software.",
    "Pipeline-security controls; artifact-signing; access model; provenance/attestation.",
    "Vendors building software for us.", "Med",
    "MAS TRM 2021 S6", "RBI IT Governance 2023 S13 (change mgmt)", "CSCRF GV.SC",
    "SOC 2", "MAS-TRM-SDLC; RBI-ITG-S13-CHANGE")
add(D, "Change management",
    "Do you operate controlled change management (testing, approval, rollback) for systems and software serving us, including vendor-managed components?",
    "Uncontrolled change is a common cause of outages and security regressions.",
    "Change-management policy; sample change records; rollback evidence.",
    "Vendors operating/changing systems for us.", "Med",
    "MAS TRM 2021 S6 (secure SDLC/change)", "RBI IT Governance 2023 S13 (change/patch)",
    "CSCRF PR", "ISO 27001 / SOC 2", "RBI-ITG-S13-CHANGE")
add(D, "Source-code escrow",
    "For critical applications, will you provide the source code to us or place it in a source-code escrow arrangement to ensure continuity if you fail or exit?",
    "RBI IT Governance requires source-code acquisition or escrow for critical applications.",
    "Escrow agreement or source-code delivery terms; release conditions.",
    "Vendors supplying critical/bespoke applications.", "Med",
    "", "RBI IT Governance 2023 S12 (source code / escrow)", "CSCRF GV.SC", "none",
    "RBI-ITG-ESCROW")

# ===========================================================================
# 9. LOGGING / SOC / INCIDENT
# ===========================================================================
D = "Logging / SOC / Incident"
add(D, "Security logging & retention",
    "Do you generate, protect and retain security/audit logs (including for systems handling our data), and what is the retention period and tamper-protection?",
    "Logs are essential for detection, investigation and regulatory evidence.",
    "Logging policy; log-source list; retention config; integrity/WORM protection.",
    "Vendors operating systems/data for us.", "Med",
    "MAS TRM 2021 S13 (logging)", "RBI CSF 2016 C16/C17; CERT-In (180-day logs in India)",
    "CSCRF DE.CM (centralised logging)", "ISO 27001 / SOC 2", "MAS-TRM-SOC; RBI-CSF-LOGS; RBI-CERTIN-LOGS")
add(D, "Log retention in India / CERT-In",
    "For India-regulated services, are ICT system logs retained for a rolling 180 days within Indian jurisdiction?",
    "CERT-In Directions require 180-day log retention within India.",
    "Log-retention attestation; storage-location evidence.",
    "India-regulated / payment / lending vendors.", "Med",
    "", "CERT-In Directions 28 Apr 2022 (180-day logs in India)", "CSCRF DE.CM", "none",
    "RBI-CERTIN-LOGS")
add(D, "24x7 monitoring / SOC",
    "Do you operate 24x7 security monitoring (SIEM/SOC, in-house or managed) with defined alerting, triage and escalation for the environment serving us?",
    "Continuous monitoring enables timely detection and response to threats.",
    "SOC operating model; SIEM use-cases; escalation matrix; SOC efficacy metrics.",
    "Material/critical vendors.", "High",
    "MAS TRM 2021 S13 (security operations/detection)", "RBI CSF 2016 C16/C17",
    "CSCRF DE.CM (24x7 SOC; M-SOC option)", "ISO 27001 / SOC 2", "MAS-TRM-SOC; SEBI-DECM-SOC")
add(D, "Log/telemetry feed to our SOC (cloud)",
    "Can you provide security logs and telemetry for ingestion and monitoring by our SOC, particularly under the cloud shared-responsibility model?",
    "RBI cloud appendix requires CSP logs to feed the RE SOC for monitoring.",
    "Log-export capability; supported formats; sample feed; shared-responsibility matrix.",
    "Cloud / hosting vendors.", "Med",
    "MAS TRM 2021 S13", "RBI IT Outsourcing 2023 Cloud (CSP logs to RE SOC)",
    "CSCRF DE.CM", "none", "RBI-ITO-CLOUDLOG; MAS-TRM-SOC")
add(D, "Incident response plan",
    "Do you maintain and test an incident-response plan covering identification, containment, eradication, recovery and reporting, including third-party-originated incidents?",
    "A tested IR plan is required to contain and recover from incidents.",
    "IR plan; named roles; last tabletop/exercise report.",
    "All vendors with access to our data/systems.", "High",
    "MAS TRM 2021 S8/S13 (incident response)", "RBI IT Governance 2023 (IR procedures)",
    "CSCRF (incident management)", "ISO 27001 / SOC 2", "MAS-TRM-INCIDENT")
add(D, "Breach notification to us",
    "Will you notify us without undue delay (within a contractually-agreed timeframe) of any incident, breach or leakage affecting our systems or data?",
    "We depend on prompt vendor notification to meet our own regulatory reporting clocks.",
    "Notification clause/SLA; notification template; named contacts; test of the channel.",
    "All vendors with access to our data/systems.", "High",
    "MAS Notice 658 (adverse-development notification)",
    "RBI Outsourcing of FS 5.6 (breach notification); RBI NBFC IT Framework 2017",
    "CSCRF GV.SC (vendor breach notification)", "none",
    "RBI-OFS-BREACH; RBI-NBFC-INCIDENT; SEBI-GVSC-BREACHNOTIFY")
add(D, "Reporting enabling 6-hour CERT-In",
    "Can you report incidents to us fast enough for us to meet our 6-hour CERT-In reporting obligation, and do you report directly to CERT-In within 6 hours where you are independently obligated?",
    "CERT-In requires reporting within 6 hours; vendors must enable the RE's clock.",
    "Incident-notification SLA aligned to 6h; CERT-In reporting procedure.",
    "India-regulated / payment / lending vendors.", "High",
    "", "CERT-In Directions 28 Apr 2022 (6-hour reporting); RBI IT Outsourcing 2023 (enable 6h RBI report)",
    "CSCRF (CERT-In 6h)", "none", "RBI-CERTIN-6H; RBI-ITO-INCIDENT6H; SEBI-INCIDENT")
add(D, "Adverse-development notification (MAS)",
    "For MAS-regulated material services, will you notify us of any adverse development (e.g., severe disruption, breach) in time for us to notify MAS within 14 working days?",
    "MAS Notice 658 requires adverse-development notification within 14 working days.",
    "Notification SLA aligned to MAS 14-working-day rule; escalation contacts.",
    "MAS material-service vendors.", "Med",
    "MAS Notice 658 (notification within 14 working days)", "", "", "none", "MAS-N658-NOTIFY")
add(D, "Root-cause analysis & lessons learned",
    "Do you provide post-incident root-cause analysis and remediation evidence to us for incidents affecting our service?",
    "RCA closes the loop and prevents recurrence; we need it for our own reporting.",
    "Sample RCA report; remediation tracking; recurrence-prevention actions.",
    "Material vendors.", "Low",
    "MAS TRM 2021 S8 (post-incident review)", "RBI IT Governance 2023 (incident analysis)",
    "CSCRF (incident reporting RCA)", "none", "MAS-TRM-INCIDENT; SEBI-INCIDENT")

# ===========================================================================
# 10. RESILIENCE & BCP
# ===========================================================================
D = "Resilience & BCP"
add(D, "BCP/DR plans",
    "Do you maintain documented, tested BCP and DR plans for the services you provide to us, aligned to our recovery objectives?",
    "Provider continuity is essential to the continuity of our critical services.",
    "BCP/DR plans; last test report; alignment to our RTO/RPO.",
    "All vendors supporting operational services.", "High",
    "MAS Guidelines on Outsourcing S5.6 (BCP); MAS BCM 2022", "RBI IT Outsourcing 2023 (BCP/DR tested)",
    "CSCRF RC.RP (recovery planning)", "ISO 27001 / ISO 22301", "MAS-OB-BCP; RBI-ITO-BCPDR; SEBI-RCRP")
add(D, "Recovery objectives (RTO/RPO/SRTO)",
    "Can you meet our required recovery-time and recovery-point objectives (and service recovery-time objective) for the services you provide?",
    "Recovery objectives quantify acceptable downtime/data-loss; vendor must commit to them.",
    "Documented RTO/RPO/SRTO commitments; evidence from recovery tests.",
    "Vendors supporting critical business services.", "High",
    "MAS BCM 2022 (SRTO per critical service)", "RBI IT Outsourcing 2023 (BCP aligned)",
    "CSCRF RC.RP (RTO/RPO)", "ISO 22301", "MAS-BCM-SRTO; SEBI-RCRP")
add(D, "BCP testing & joint exercises",
    "Do you test BCP/DR at least annually and participate in joint continuity exercises with us where we support a critical service?",
    "Untested plans fail; joint exercises validate end-to-end recovery.",
    "Test schedule; latest test results; willingness for joint exercises.",
    "Critical-service vendors.", "Med",
    "MAS BCM 2022 (third-party BCP testing)", "RBI IT Outsourcing 2023 (periodic BCP testing)",
    "CSCRF RC.RP", "ISO 22301", "MAS-BCM-TEST; RBI-ITO-BCPDR")
add(D, "Contractual recovery commitments",
    "Are your continuity/recovery obligations (including for critical services) embedded as contractual commitments consistent with our objectives?",
    "MAS BCM requires recovery obligations for critical services to be contractual.",
    "Contract clauses on recovery/continuity; SLA-linked recovery commitments.",
    "Critical-service vendors.", "Med",
    "MAS BCM 2022 (contractual recovery commitments)", "RBI IT Outsourcing 2023", "CSCRF RC.RP",
    "none", "MAS-BCM-CONTRACT")
add(D, "Dependency mapping",
    "Have you mapped the dependencies (sub-contractors, infrastructure, locations) underpinning the critical services you provide to us, identifying single points of failure?",
    "Dependency/SPOF mapping is required to understand and manage resilience risk.",
    "Dependency map for our critical services; SPOF analysis; mitigation plan.",
    "Critical-service vendors.", "Med",
    "MAS BCM 2022 (third-party dependency mapping)", "RBI IT Governance 2023 S10 (single point of failure)",
    "CSCRF GV.SC", "none", "MAS-BCM-DEPENDENCY; RBI-ITG-S10")
add(D, "Resilience concentration risk",
    "Do multiple of our critical services rely on the same of your facilities, infrastructure or sub-contractors, and how is that concentration mitigated?",
    "Concentration creates correlated failure risk across our critical services.",
    "Concentration analysis; geographic/infrastructure diversity; mitigation options.",
    "Critical-service vendors.", "Med",
    "MAS BCM 2022 (resilience concentration)", "RBI IT Outsourcing 2023 (concentration); RBI Outsourcing of FS 5.4",
    "CSCRF GV.SC", "none", "MAS-BCM-CONCENTRATION; RBI-ITO-CONCENTRATION; RBI-OFS-CONCENTRATION")
add(D, "Backup & restoration",
    "How do you back up our data, where are backups stored (residency), are they encrypted and immutable, and how often are restorations tested?",
    "Reliable, tested, protected backups are the last line against data loss and ransomware.",
    "Backup policy; encryption/immutability; restoration-test evidence; backup residency.",
    "Vendors storing our data.", "Med",
    "MAS TRM 2021 (availability/recovery)", "RBI IT Outsourcing 2023 (BCP/DR)", "CSCRF RC.RP",
    "ISO 27001 (A.8.13)", "RBI-ITO-BCPDR; SEBI-RCRP")

# ===========================================================================
# 11. CYBER AUDIT / VAPT / ASSURANCE
# ===========================================================================
D = "Cyber Audit / VAPT"
add(D, "Independent security certification",
    "Do you hold a current independent security certification/attestation (ISO/IEC 27001, SOC 2 Type II, PCI DSS as applicable) covering the service you provide to us?",
    "Independent assurance reduces our assessment effort and evidences control maturity.",
    "Certificate + statement of applicability/scope; SOC 2 report; PCI AOC.",
    "All material vendors.", "High",
    "MAS Guidelines on Outsourcing S5.9 (independent assessment)", "RBI IT Outsourcing 2023 (assurance)",
    "CSCRF ISO 27001 (MIIs/critical facilities)", "ISO 27001 / SOC 2 / PCI", "MAS-OB-AUDIT3Y; SEBI-ISO27001")
add(D, "ISO 27001 for critical facilities",
    "For data centres, DR/NDR sites, SOC or colocation supporting our SEBI-regulated service, do those facilities hold ISO/IEC 27001?",
    "CSCRF mandates ISO 27001 for MIIs and critical third-party facilities.",
    "ISO 27001 certificate scoped to the relevant facility.",
    "SEBI-regulated critical facilities.", "Med",
    "", "", "CSCRF ISO 27001 mandate (PDC/DR/NDR, SOC, DC/colocation)", "ISO 27001", "SEBI-ISO27001")
add(D, "Independent audit cadence",
    "Is the service subject to independent audit / expert control assessment at least every 3 years (or per our materiality), with reports shared with us?",
    "MAS requires independent audit of material outsourcing at least every 3 years.",
    "Audit schedule; latest independent audit/assurance report.",
    "Material vendors.", "Med",
    "MAS Guidelines on Outsourcing S5.9 (audit <=3 years)", "RBI IT Governance 2023 (IS audit covers third parties)",
    "CSCRF VAPT/Cyber Audit", "ISO 27001 / SOC 2", "MAS-OB-AUDIT3Y; RBI-ITG-ISAUDIT")
add(D, "VAPT cadence",
    "Do you conduct regular VAPT (incl. internet-facing/third-party-exposed systems serving us) at a defined cadence, and remediate findings risk-based?",
    "Regular VAPT validates control effectiveness against real attack techniques.",
    "VAPT schedule; latest executive summary; remediation status of findings.",
    "Internet-facing / critical vendors.", "High",
    "MAS TRM 2021 S14 (VAPT)", "RBI IT Governance 2023 (annual PT, semi-annual VA); RBI CSF C7",
    "CSCRF VAPT (cadence by RE category)", "ISO 27001 / SOC 2 / PCI", "MAS-TRM-VAPT; SEBI-VAPT; RBI-CSF-PATCHVULN")
add(D, "CERT-In empanelled auditor",
    "For India-regulated services, are your VAPT/cyber audits and (for payment data) System Audit Report performed by a CERT-In empanelled auditor?",
    "RBI/SEBI require cyber audits/SAR by CERT-In empanelled auditors.",
    "Auditor empanelment evidence; latest SAR/cyber-audit report.",
    "India-regulated / payment vendors.", "High",
    "", "RBI Storage of Payment System Data 2018 (CERT-In SAR); RBI IT Governance 2023",
    "CSCRF VAPT/Cyber Audit (CERT-In empanelled)", "none", "RBI-PSD-SAR; SEBI-VAPT")
add(D, "Adversarial / red-team testing",
    "Do you (or we, with your cooperation) conduct scenario-based red-team / attack-simulation exercises against the service to validate detection and response?",
    "Red-teaming tests real-world detection/response beyond checklist VAPT.",
    "Red-team scope/cadence; summary of last exercise; remediation actions.",
    "High-criticality vendors.", "Low",
    "MAS TRM 2021 S14 (adversarial attack simulation)", "RBI IT Governance 2023", "CSCRF VAPT",
    "none", "MAS-TRM-REDTEAM")
add(D, "IT audit covers third parties",
    "Do you cooperate with our IS/IT audit (and our regulator's) of outsourced/third-party services, providing access and timely responses?",
    "IS audit must cover outsourced services with right to audit the provider.",
    "Audit-cooperation commitment; evidence of past audit support.",
    "Material vendors.", "Med",
    "MAS TRM 2021 S4 (IT audit)", "RBI IT Governance 2023 (IS audit covers third parties); RBI IT Outsourcing 2023",
    "CSCRF GV.SC", "none", "MAS-TRM-ITAUDIT; RBI-ITG-ISAUDIT")

# ===========================================================================
# 12. PRIVACY
# ===========================================================================
D = "Privacy"
add(D, "Privacy policy & lawful basis",
    "Do you maintain a privacy policy and process personal data we share only on documented instructions and a lawful basis, acting as processor where applicable?",
    "Ensures lawful, instruction-bound processing of personal data we control.",
    "Privacy policy; data-processing agreement; processor-role acknowledgement.",
    "Vendors processing personal data.", "Med",
    "PDPA 2012 (Singapore)", "RBI Digital Lending 2025 (data privacy); DPDP Act 2023 (India)",
    "CSCRF PR.DS", "ISO 27701", "MAS-PDPA-XBORDER; RBI-DL-DATAMIN")
add(D, "Consent management",
    "How do you capture, record and honour customer consent for collection, use and sharing of personal data, including withdrawal?",
    "Consent-driven processing is required, especially in digital lending and customer-data sharing.",
    "Consent-capture mechanism; consent records; withdrawal handling.",
    "Customer-facing / lending vendors.", "Med",
    "MAS Guidelines on Outsourcing (customer-info consent/disclosure)",
    "RBI Digital Lending 2025 (consent-driven data)", "CSCRF GV.SC", "ISO 27701",
    "MAS-OB-CONSENT; RBI-DL-DATAMIN")
add(D, "Grievance redressal officer (lending)",
    "As an LSP/DLA, have you designated a nodal grievance-redressal officer, and is a complaint facility available on the app/website?",
    "Digital-lending rules require a designated grievance officer at the LSP/DLA.",
    "GRO appointment; complaint-channel screenshot; escalation to ombudsman path.",
    "LSPs / DLAs.", "Med",
    "", "RBI Digital Lending 2025 (grievance officer); RBI Outsourcing of FS 5.7",
    "", "none", "RBI-DL-GRO; RBI-OFS-GRIEVANCE")
add(D, "Privacy-by-design & PII handling",
    "Do you apply data-minimisation, purpose-limitation and PII-handling controls (masking, tokenisation, restricted access) by design?",
    "Reduces privacy risk surface and limits exposure of personal data.",
    "Privacy-by-design evidence; PII-masking/tokenisation; access restrictions.",
    "Vendors processing PII.", "Low",
    "PDPA 2012", "RBI Digital Lending 2025 (minimisation); DPDP Act 2023", "CSCRF PR.DS", "ISO 27701",
    "RBI-DL-DATAMIN")
add(D, "Breach notification (privacy)",
    "Will you notify us of any personal-data breach in time for us to meet our data-protection notification obligations to regulators and data principals?",
    "Privacy regimes impose breach-notification duties dependent on processor reporting.",
    "Privacy-breach notification SLA; template; named DPO/contact.",
    "Vendors processing personal data.", "Med",
    "PDPA 2012 (breach notification)", "DPDP Act 2023 (breach notification)", "CSCRF GV.SC",
    "ISO 27701", "MAS-PDPA-XBORDER; SEBI-GVSC-BREACHNOTIFY")

# ===========================================================================
# 13. SUB-CONTRACTING / 4TH PARTY
# ===========================================================================
D = "Sub-contracting / 4th-party"
add(D, "Sub-contractor disclosure",
    "Will you disclose and keep current the list of sub-contractors (4th parties) involved in delivering our service, including their role and location?",
    "We must understand the full chain to manage 4th-party and concentration risk.",
    "Sub-processor list; roles/locations; change-notification process.",
    "Vendors using sub-contractors.", "High",
    "MAS Guidelines on Outsourcing S5.7 (sub-contracting)", "RBI IT Outsourcing 2023 (sub-contracting)",
    "CSCRF GV.SC (material subcontractors)", "none", "MAS-OB-SUBCON; RBI-ITO-SUBCON; SEBI-GVSC-SUBAUDIT")
add(D, "Prior consent for material sub-contracting",
    "Do you obtain our prior written consent (or give prior notice) before engaging or materially changing a sub-contractor handling our data/critical functions?",
    "Material sub-contracting requires the institution's prior knowledge/consent.",
    "Sub-contracting clause requiring consent/notice; sample consent record.",
    "Vendors using material sub-contractors.", "High",
    "MAS Guidelines on Outsourcing S5.7 (prior consent)", "RBI IT Outsourcing 2023 (prior consent; provider liable)",
    "CSCRF GV.SC", "none", "MAS-OB-SUBCON; RBI-ITO-SUBCON")
add(D, "Provider remains liable for sub-contractors",
    "Do you remain fully accountable and liable for the acts, omissions and security posture of your sub-contractors handling our data/service?",
    "Accountability cannot be passed down the chain; the prime provider stays liable.",
    "Liability/flow-down clause; sub-contractor accountability statement.",
    "Vendors using sub-contractors.", "High",
    "MAS Guidelines on Outsourcing S5.7", "RBI IT Outsourcing 2023 (provider fully liable)",
    "CSCRF GV.SC (RE accountable; vendor flows down)", "none", "MAS-OB-SUBCON; RBI-ITO-SUBCON")
add(D, "Flow-down of controls & audit rights",
    "Do you contractually flow down equivalent security, data-protection, audit and breach-notification obligations to material sub-contractors, within our audit scope?",
    "Controls and audit rights must reach material 4th parties to be effective.",
    "Flow-down clauses; sub-contractor audit-rights confirmation.",
    "Vendors using material sub-contractors.", "High",
    "MAS Guidelines on Outsourcing S5.7 (equivalent controls)", "RBI IT Outsourcing 2023 (audit of sub-contractors)",
    "CSCRF GV.SC (audit of material subcontractors)", "none",
    "MAS-OB-SUBCON; RBI-ITO-AUDIT; SEBI-GVSC-SUBAUDIT")
add(D, "Sub-contractor due diligence",
    "Do you perform security due diligence and ongoing assessment on sub-contractors handling our data, equivalent to what we expect of you?",
    "4th-party risk must be assessed, not assumed, by the prime provider.",
    "Sub-contractor DD process; sample assessment; cadence.",
    "Vendors using sub-contractors.", "Med",
    "MAS Guidelines on Outsourcing S5.7", "RBI IT Outsourcing 2023 (DD on sub-contractors)",
    "CSCRF GV.SC", "none", "MAS-OB-SUBCON; RBI-ITO-DD-STAFF; SEBI-GVSC-SUBAUDIT")
add(D, "Sub-contractor concentration",
    "Is there concentration in your sub-contractor base (e.g., a single shared cloud/SaaS underpinning many functions) that could create correlated failure?",
    "4th-party concentration can undermine resilience despite vendor diversity.",
    "4th-party concentration analysis; critical-dependency identification.",
    "Vendors with critical sub-contractors.", "Low",
    "MAS BCM 2022 (concentration)", "RBI IT Outsourcing 2023 (concentration)", "CSCRF GV.SC",
    "none", "MAS-BCM-CONCENTRATION; RBI-ITO-CONCENTRATION")

# ===========================================================================
# 14. PAYMENTS-SPECIFIC (PSS / PA-PG / TOKENIZATION / DL)
# ===========================================================================
D = "Payments-specific"
add(D, "PA/PG authorisation & standards",
    "As a payment aggregator/gateway, are you RBI-authorised (or operating under a valid timeline) and compliant with the PA/PG technology, security and merchant-onboarding requirements?",
    "PA/PG activities require RBI authorisation and adherence to prescribed standards.",
    "RBI PA authorisation/in-principle; PA/PG compliance attestation.",
    "Payment aggregators/gateways.", "High",
    "", "RBI PA/PG Guidelines 2020; RBI PA Master Directions 2025 (authorisation/standards)",
    "", "PCI DSS", "RBI-PSD-LOC")
add(D, "Merchant onboarding / CDD",
    "Do you perform customer due diligence (CDD/KYC) on merchants you onboard on our behalf, using valid documents and risk-based onboarding?",
    "PA Directions require comprehensive merchant CDD to prevent misuse.",
    "Merchant-onboarding/CDD policy; KYC records; risk-tiering of merchants.",
    "Payment aggregators.", "High",
    "", "RBI PA/PG Guidelines 2020; RBI PA Master Directions 2025 (merchant onboarding/CDD)",
    "", "none", "RBI-PSD-LOC")
add(D, "Escrow / settlement controls",
    "For non-bank PA activity, are merchant funds held in escrow with a scheduled commercial bank, with required settlement and (for cross-border) InCA/OCA controls?",
    "PA Directions mandate escrow-based fund handling and segregated cross-border accounts.",
    "Escrow-account confirmation; settlement-cycle evidence; InCA/OCA setup if applicable.",
    "Non-bank payment aggregators.", "Med",
    "", "RBI PA/PG Guidelines 2020; RBI PA Master Directions 2025 (escrow; InCA/OCA)", "",
    "none", "RBI-PSD-LOC")
add(D, "No Card-on-File storage (CoFT)",
    "Do you confirm that no actual card data (PAN/CVV/expiry) is stored by you or your systems, with tokenisation used and any previously stored CoF data purged?",
    "Only card issuers/networks may store CoF data; others must tokenise and purge.",
    "Confirmation of no CoF storage; tokenisation architecture; purge evidence.",
    "Entities in the card payment chain.", "High",
    "", "RBI Card-on-File Tokenisation (effective 1 Jan 2022 — no CoF storage); RBI PA/PG Guidelines",
    "", "PCI DSS", "RBI-PSD-LOC")
add(D, "Tokenisation controls",
    "Where you tokenise card data, are tokens unique per card+token-requestor+merchant, generated with explicit customer consent and AFA, and is token-vault access tightly controlled?",
    "CoFT requires consent+AFA and uniqueness constraints to limit token misuse.",
    "Tokenisation design; consent/AFA flow; token-vault access controls.",
    "Token requestors / tokenisation providers.", "High",
    "", "RBI Card-on-File Tokenisation (token uniqueness; consent + AFA)", "", "PCI DSS",
    "RBI-PSD-LOC")
add(D, "PCI DSS compliance",
    "Are you PCI DSS (and PA-DSS where applicable) compliant for the cardholder-data environment supporting our service, with a current AOC/ROC?",
    "PCI DSS is the baseline for cardholder-data protection in the payment chain.",
    "PCI DSS AOC/ROC; scope/SAQ; remediation status.",
    "Vendors handling cardholder data.", "High",
    "", "RBI PA/PG Guidelines (PCI DSS/PA-DSS adherence)", "", "PCI DSS", "RBI-PSD-LOC")
add(D, "Payment-data System Audit Report",
    "Do you submit (and can you share) a System Audit Report by a CERT-In empanelled auditor confirming payment-data localisation and security, board-approved?",
    "RBI payment-data localisation compliance is evidenced via a CERT-In SAR.",
    "Latest SAR; board-approval record; submission evidence to RBI.",
    "Payment system providers.", "High",
    "", "RBI Storage of Payment System Data 2018 (CERT-In SAR)", "", "none", "RBI-PSD-SAR; RBI-PSD-LOC")
add(D, "Digital-lending app controls (DLA)",
    "For digital-lending apps, do you ensure transparency (KFS, fees, RE identity), no automatic credit-limit increases without consent, and compliant data access (no contacts/media/files harvesting)?",
    "Digital-lending rules constrain DLA conduct and on-device data access.",
    "DLA permission inventory; consent flows; KFS/disclosure screenshots.",
    "LSPs / DLAs.", "High",
    "", "RBI Digital Lending 2025 (DLA conduct/transparency/data access)", "", "none",
    "RBI-DL-DD; RBI-DL-DATAMIN")

# ===========================================================================
# 15. CLOUD-SPECIFIC
# ===========================================================================
D = "Cloud-specific"
add(D, "Shared-responsibility model",
    "Do you provide a clear shared-responsibility matrix delineating security responsibilities between you (CSP) and us across IaaS/PaaS/SaaS layers?",
    "Misunderstood responsibility boundaries are a leading cloud-risk cause.",
    "Shared-responsibility matrix mapped to our controls.",
    "Cloud / hosting vendors.", "High",
    "MAS TRM 2021 (cloud/third-party)", "RBI IT Outsourcing 2023 Appendix I Cloud (shared responsibility)",
    "CSCRF Cloud", "ISO 27017 / SOC 2", "RBI-ITO-CLOUDPATCH; SEBI-CLOUD")
add(D, "Cloud empanelment / localisation (SEBI)",
    "For SEBI-regulated services, is cloud provided via a MeitY-empanelled and STQC-certified CSP, with market-critical/regulatory data stored and processed within India?",
    "CSCRF cloud framework requires empanelled/STQC CSPs and India localisation for critical data.",
    "MeitY empanelment + STQC certification evidence; data-region attestation.",
    "SEBI-regulated cloud services.", "High",
    "", "", "CSCRF Cloud; SEBI Cloud Framework (MeitY/STQC; India localisation)", "none", "SEBI-CLOUD")
add(D, "Cloud configuration & CSPM",
    "How do you prevent and detect cloud misconfiguration (e.g., public storage, over-permissive IAM) — do you operate CSPM and enforce secure baselines?",
    "Misconfiguration is the dominant cause of cloud data exposure.",
    "CSPM tooling; baseline policies; sample misconfiguration findings/remediation.",
    "Cloud-hosted vendors.", "Med",
    "MAS TRM 2021 S12 (infrastructure security)", "RBI IT Outsourcing 2023 Cloud (secure config)",
    "CSCRF Cloud / PR", "ISO 27017 / SOC 2", "MAS-TRM-DATASEC; RBI-CSF-SECCONFIG")
add(D, "Cloud IAM (RBAC/MFA)",
    "Do you enforce RBAC, least-privilege and MFA for privileged and remote access to the cloud environment hosting our data?",
    "RBI cloud appendix explicitly requires RBAC, least-privilege and MFA.",
    "Cloud IAM policy; MFA coverage; privileged-role inventory.",
    "Cloud-hosted vendors.", "High",
    "MAS TRM 2021 S9", "RBI IT Outsourcing 2023 Cloud (RBAC/least-privilege/MFA)",
    "CSCRF PR.AA", "ISO 27017 / SOC 2", "RBI-ITO-CLOUDRBAC; SEBI-PRAA-MFA")
add(D, "Cloud exit / portability",
    "Can we exit the cloud arrangement in an orderly way with data portability/return, no lock-in barriers, and continuity during transition?",
    "Cloud lock-in undermines exitability; portability must be designed in.",
    "Exit/portability plan; data-export formats; transition support terms.",
    "Cloud-hosted vendors.", "Med",
    "MAS Guidelines on Outsourcing S5.9 (exit)", "RBI IT Outsourcing 2023 (exit; cloud)",
    "CSCRF Cloud / GV.SC", "none", "MAS-OB-TERMINATION; RBI-ITO-EXIT")
add(D, "Cloud resilience & multi-AZ",
    "How is resilience achieved in the cloud (multi-AZ/region, automated failover) for the services hosting our critical operations?",
    "Cloud resilience design determines whether our recovery objectives are achievable.",
    "Architecture showing AZ/region redundancy; failover test results.",
    "Cloud-hosted critical services.", "Med",
    "MAS BCM 2022", "RBI IT Outsourcing 2023 (BCP/DR; cloud)", "CSCRF RC.RP", "ISO 22301",
    "SEBI-RCRP; RBI-ITO-BCPDR")

# ===========================================================================
# 16. EXIT & TERMINATION (cross-cutting; small dedicated group folded into Governance)
# ===========================================================================
D = "Governance & Outsourcing"
add(D, "Exit strategy",
    "Is there a documented exit/contingency strategy enabling orderly transfer or in-sourcing, data return/destruction, and continuity during transition?",
    "A pre-agreed exit prevents disruption and data loss if the arrangement ends.",
    "Exit plan; transition timeline; data-return/destruction terms.",
    "Material vendors.", "High",
    "MAS Guidelines on Outsourcing S5.9 (exit strategy)", "RBI IT Outsourcing 2023 (exit; no data erasure during transition)",
    "CSCRF GV.SC (secure exit)", "none", "MAS-OB-TERMINATION; RBI-ITO-EXIT; SEBI-GVSC-CONTRACT")
add(D, "No data withholding on exit",
    "Will you support an orderly exit, return our data, and refrain from deleting or withholding our data during the transition period?",
    "RBI requires that providers not delete/withhold RE data during transition.",
    "Contractual no-withholding clause; data-return runbook.",
    "Material vendors holding our data.", "High",
    "MAS Guidelines on Outsourcing S5.9", "RBI IT Outsourcing 2023 (no data erasure during transition)",
    "CSCRF GV.SC", "none", "RBI-ITO-EXIT")
add(D, "Vendor support / maintenance agreement",
    "For acquired/critical applications, is there a formal support and maintenance agreement covering the full lifecycle (updates, security fixes, EoL notice)?",
    "Lifecycle support is required to keep critical applications secure and available.",
    "Support/maintenance agreement; SLA; EoL-notice commitment.",
    "Vendors supplying critical applications.", "Med",
    "MAS Guidelines on Outsourcing", "RBI IT Governance 2023 S12 (vendor support agreement)",
    "CSCRF GV.SC", "none", "RBI-ITG-S12-SUPPORT")
add(D, "Concentration risk (provider)",
    "Are we over-reliant on you for multiple material services, and what alternatives/portability exist to mitigate provider concentration?",
    "Provider concentration is a systemic and operational risk requiring active management.",
    "Service-dependency overview; substitutability assessment.",
    "Material multi-service vendors.", "Med",
    "MAS Guidelines on Outsourcing (concentration)",
    "RBI IT Outsourcing 2023 (concentration); RBI Outsourcing of FS 5.4", "CSCRF GV.SC",
    "none", "MAS-OB-CONCENTRATION; RBI-ITO-CONCENTRATION; RBI-OFS-CONCENTRATION")

# ===========================================================================
# Generic baseline domain (regulator-silent areas) — ensures "None"/custom vendors
# still receive a solid set.
# ===========================================================================
D = "Security Baseline (general)"
add(D, "ISMS / security policy framework",
    "Do you maintain a documented information-security management system / policy framework, approved and reviewed at least annually?",
    "A governing policy framework is the foundation of a defensible security programme.",
    "ISMS scope; policy index; last review/approval record.",
    "All vendors.", "Med",
    "MAS TRM 2021 S3 (governance)", "RBI IT Governance 2023", "CSCRF GV", "ISO 27001 (A.5.1)",
    "MAS-TRM-GOV")
add(D, "Risk assessment process",
    "Do you operate a documented information-security risk-assessment process feeding treatment decisions and management reporting?",
    "Risk-based prioritisation requires a repeatable assessment process.",
    "Risk-assessment methodology; current risk register (sanitised).",
    "All vendors.", "Low",
    "MAS TRM 2021 S3", "RBI IT Governance 2023", "CSCRF GV/ID", "ISO 27001 (A.5)",
    "MAS-TRM-GOV")
add(D, "Asset management",
    "Do you maintain an inventory of information assets (hardware, software, data) supporting our service, with assigned owners?",
    "Asset visibility underpins patching, hardening and incident response.",
    "Asset inventory excerpt; ownership model.",
    "All vendors.", "Low",
    "MAS TRM 2021 S12 (asset inventory)", "RBI CSF 2016 C5", "CSCRF ID", "ISO 27001 (A.5.9)",
    "MAS-TRM-DATASEC")
add(D, "Physical & environmental security",
    "How do you control physical access to facilities/data centres housing our data, and what environmental controls protect availability?",
    "Physical access and environmental failures can compromise confidentiality and availability.",
    "Physical-access controls; data-centre certifications; environmental controls.",
    "Vendors operating physical facilities/DCs.", "Low",
    "MAS TRM 2021 (infrastructure)", "RBI CSF 2016", "CSCRF PR", "ISO 27001 (A.7)",
    "MAS-TRM-DATASEC")
add(D, "Vulnerability disclosure / bug bounty",
    "Do you operate a vulnerability-disclosure programme or bug bounty and a security point of contact for reporting issues affecting our service?",
    "A disclosure channel surfaces externally-found vulnerabilities responsibly.",
    "VDP/bug-bounty policy; security.txt / contact; intake-to-fix process.",
    "Internet-facing vendors.", "Low",
    "MAS TRM 2021 S14", "RBI IT Governance 2023", "CSCRF DE", "none", "MAS-TRM-VAPT")
add(D, "Insurance / cyber cover",
    "Do you maintain adequate cyber-liability and professional-indemnity insurance covering incidents affecting our service?",
    "Insurance backs the vendor's liability and breach-cost obligations.",
    "Insurance certificates with coverage limits and scope.",
    "Material vendors.", "Low",
    "MAS Guidelines on Outsourcing (financial standing)", "RBI IT Outsourcing 2023", "CSCRF GV.SC",
    "none", "MAS-OB-DUEDIL")

# ===========================================================================
# SUPPLEMENTARY DEPTH — additional non-duplicative controls to broaden coverage
# ===========================================================================
D = "Cryptography & Key Mgmt"
add(D, "Hashing & password storage",
    "How are passwords and other sensitive credentials hashed/salted (e.g., bcrypt/scrypt/Argon2), and do you avoid weak hashes (MD5/SHA-1) anywhere in scope?",
    "Weak or unsalted hashing enables credential cracking after a breach.",
    "Hashing standard; algorithm/parameters; evidence weak hashes are not in use.",
    "Vendors storing credentials/authenticators.", "Med",
    "MAS TRM 2021 S12 (cryptography)", "RBI CSF 2016 C8", "CSCRF PR.DS", "ISO 27001", "MAS-TRM-CRYPTO")
add(D, "Crypto-agility & PQC readiness",
    "Can you upgrade cryptographic algorithms without major re-architecture, and do you have a roadmap toward post-quantum-ready cryptography?",
    "Crypto-agility limits future risk from algorithm deprecation/compromise.",
    "Crypto-agility statement; algorithm-inventory; PQC roadmap if any.",
    "Long-lived/material vendors.", "Low",
    "MAS TRM 2021 S12", "RBI IT Governance 2023", "CSCRF PR.DS", "none", "MAS-TRM-CRYPTO")

D = "Access Control & MFA"
add(D, "Service / machine account control",
    "How are service, API and machine accounts inventoried, secured, rotated and prevented from interactive login?",
    "Unmanaged service accounts are a common privilege-escalation path.",
    "Service-account inventory; rotation policy; interactive-login restrictions.",
    "Vendors with automated integrations.", "Med",
    "MAS TRM 2021 S9", "RBI CSF 2016 C8", "CSCRF PR.AA", "ISO 27001 / SOC 2", "MAS-TRM-ACCESS")
add(D, "Phishing-resistant / strong MFA",
    "For high-risk access, do you use phishing-resistant MFA (FIDO2/hardware tokens) rather than SMS/OTP alone?",
    "SMS/OTP MFA is bypassable; phishing-resistant factors materially reduce account-takeover risk.",
    "MFA factor types in use; coverage for privileged/high-risk access.",
    "Vendors with privileged access to critical systems.", "Low",
    "MAS Notice 655 4.6 (MFA)", "RBI CSF 2016 C8", "CSCRF PR.AA (MFA, zero-trust aligned)",
    "ISO 27001", "MAS-655-MFA; SEBI-PRAA-MFA")

D = "Vulnerability & Patch"
add(D, "Web/application vulnerability coverage",
    "Do your security tests cover application-layer vulnerabilities (OWASP Top 10: injection, XSS, broken access control, SSRF, deserialization) for the service we consume?",
    "Application-layer flaws are a leading breach cause and are not caught by infra scanning alone.",
    "App pen-test/DAST scope and results; OWASP coverage mapping.",
    "Vendors providing web/app/API services.", "Med",
    "MAS TRM 2021 S14 (VAPT)", "RBI CSF 2016 6.2 (app security)", "CSCRF VAPT",
    "ISO 27001 / SOC 2 / PCI", "MAS-TRM-VAPT; RBI-CSF-APPSEC")
add(D, "Container / image vulnerability mgmt",
    "If you run containers, how are base images hardened and scanned for vulnerabilities, and how is image provenance/registry access controlled?",
    "Vulnerable or untrusted container images propagate risk across the platform.",
    "Image-scanning evidence; hardened base-image policy; registry access controls.",
    "Vendors using containers/Kubernetes.", "Low",
    "MAS TRM 2021 S12", "RBI IT Outsourcing 2023 Cloud (patch)", "CSCRF PR.MA",
    "ISO 27001", "MAS-TRM-PATCH; RBI-ITO-CLOUDPATCH")

D = "Privacy"
add(D, "Data-subject rights handling",
    "How do you support data-subject/customer rights (access, correction, erasure) for personal data you process on our behalf, within statutory timelines?",
    "We rely on the processor to action rights requests we receive from data principals.",
    "Rights-handling procedure; SLA; sample fulfilment record.",
    "Vendors processing personal data.", "Low",
    "PDPA 2012 (access/correction)", "DPDP Act 2023 (data-principal rights)", "CSCRF PR.DS",
    "ISO 27701", "MAS-PDPA-XBORDER")
add(D, "Purpose limitation on data sharing",
    "Do you ensure our customers' data is used only for the agreed purpose and never sold, shared, or used for your own profiling/marketing without our authorisation?",
    "Prevents secondary use and unauthorised onward sharing of customer data.",
    "Contractual purpose-limitation; no-secondary-use attestation; data-sharing log.",
    "Vendors processing customer/personal data.", "Med",
    "MAS Guidelines on Outsourcing (customer-info use)", "RBI Digital Lending 2025 (purpose limitation)",
    "CSCRF GV.SC", "ISO 27701", "MAS-OB-CONSENT; RBI-DL-DATAMIN")

D = "Due Diligence & Contracts"
add(D, "Regulatory-change cooperation",
    "Will you cooperate to implement new regulatory requirements affecting our service during the contract term, within agreed timelines and without disproportionate cost?",
    "Regulations evolve; the contract must allow us to remain compliant via the vendor.",
    "Change-cooperation clause; example of past regulatory uplift delivered.",
    "Material vendors.", "Low",
    "MAS Guidelines on Outsourcing (ongoing compliance)", "RBI IT Outsourcing 2023 (compliance)",
    "CSCRF GV.SC", "none", "MAS-OB-AGREEMENT; SEBI-GVSC-CONTRACT")

D = "Logging / SOC / Incident"
add(D, "Forensic readiness & evidence preservation",
    "Can you preserve forensic evidence (logs, images, artefacts) and support our investigation/regulatory enquiry following an incident affecting our service?",
    "Forensic readiness is essential for root-cause analysis and regulatory defensibility.",
    "Forensic-readiness procedure; evidence-handling/chain-of-custody; retention for investigations.",
    "Material vendors.", "Low",
    "MAS TRM 2021 S8/S13", "RBI IT Governance 2023 (incident analysis)", "CSCRF (incident management)",
    "ISO 27001 / SOC 2", "MAS-TRM-INCIDENT; SEBI-INCIDENT")

D = "Cloud-specific"
add(D, "Tenant key & encryption (cloud SaaS)",
    "In a SaaS model, is our tenant data encrypted with per-tenant keys (or BYOK), and can we revoke access cryptographically on exit?",
    "Per-tenant keying strengthens isolation and supports crypto-erasure on exit.",
    "Per-tenant/BYOK key model; crypto-shredding capability on exit.",
    "Multi-tenant SaaS vendors.", "Med",
    "MAS TRM 2021 S12", "RBI IT Outsourcing 2023 Cloud (keys under RE control)", "CSCRF PR.DS",
    "ISO 27017", "RBI-ITO-CLOUDKEY; SEBI-PRDS-KEYLOC")

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------
# Domain ID prefixes for stable Control IDs
DOMAIN_PREFIX = OrderedDict([
    ("Governance & Outsourcing", "GOV"),
    ("Due Diligence & Contracts", "DDC"),
    ("Data Security & Residency", "DSR"),
    ("Cryptography & Key Mgmt", "CKM"),
    ("Access Control & MFA", "ACM"),
    ("Network & Malware", "NWM"),
    ("Vulnerability & Patch", "VPM"),
    ("SDLC & Supply Chain (SBOM)", "SDL"),
    ("Logging / SOC / Incident", "LSI"),
    ("Resilience & BCP", "RBC"),
    ("Cyber Audit / VAPT", "AUD"),
    ("Privacy", "PRV"),
    ("Sub-contracting / 4th-party", "SUB"),
    ("Payments-specific", "PAY"),
    ("Cloud-specific", "CLD"),
    ("Security Baseline (general)", "BAS"),
])

DOMAIN_DESC = {
    "Governance & Outsourcing": "Board accountability, outsourcing register, materiality, exit and concentration.",
    "Due Diligence & Contracts": "Onboarding/periodic due diligence, contractual terms, audit/inspection rights, SLAs.",
    "Data Security & Residency": "Classification, encryption, data localisation (India/SEBI keys), minimisation, segregation, disposal.",
    "Cryptography & Key Mgmt": "Crypto standards, key lifecycle, customer key control, certificates and secrets.",
    "Access Control & MFA": "Least privilege/RBAC, MFA, PAM, admin accounts, JML and segregation of duties.",
    "Network & Malware": "Perimeter/segmentation, malware/EDR, hardening, secure API connectivity and remote access.",
    "Vulnerability & Patch": "Patch process and SLAs, vulnerability scanning, threat intel and EoL management.",
    "SDLC & Supply Chain (SBOM)": "Secure SDLC, code review, clean-code certification, SBOM, dependency and pipeline security, escrow.",
    "Logging / SOC / Incident": "Logging/retention, 24x7 SOC, incident response, breach notification (CERT-In 6h, MAS 14 wd).",
    "Resilience & BCP": "BCP/DR, RTO/RPO/SRTO, testing, dependency mapping, concentration and backups.",
    "Cyber Audit / VAPT": "Independent certifications, audit cadence, VAPT, CERT-In empanelled audit and red-teaming.",
    "Privacy": "Privacy policy/lawful basis, consent, grievance officer, privacy-by-design and breach notification.",
    "Sub-contracting / 4th-party": "Sub-contractor disclosure, prior consent, liability flow-down, DD and concentration.",
    "Payments-specific": "PA/PG, merchant onboarding, escrow, CoFT/tokenisation, PCI DSS, payment SAR and DLA controls.",
    "Cloud-specific": "Shared responsibility, empanelment/localisation, configuration/CSPM, IAM, exit and resilience.",
    "Security Baseline (general)": "Regulator-agnostic baseline for any vendor: ISMS, risk, assets, physical, VDP, insurance.",
}

FRAMEWORKS = [
    ("MAS", 8),   # column index in tuple (0-based) for mapping
    ("RBI", 9),
    ("SEBI", 10),
]

# Styling helpers
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DOMAIN_FILL = PatternFill("solid", fgColor="D9E1F2")
DOMAIN_FONT = Font(bold=True, size=11, color="1F3864")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
CENTER = Alignment(horizontal="center", vertical="top")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

WEIGHT_FILL = {
    "High": PatternFill("solid", fgColor="F8CBAD"),
    "Med": PatternFill("solid", fgColor="FFE699"),
    "Low": PatternFill("solid", fgColor="C6E0B4"),
}

wb = Workbook()

# ---- Sheet 1: Controls ----
ws = wb.active
ws.title = "Controls"
ws.append(HEADERS)
for c in ws[1]:
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDER

# Order controls by the domain order in DOMAIN_PREFIX, preserving insertion order within each.
controls_by_domain = OrderedDict((d, []) for d in DOMAIN_PREFIX)
for row in C:
    controls_by_domain[row[0]].append(row)

domain_counts = Counter()
framework_domain_counts = {fw: Counter() for fw, _ in FRAMEWORKS}
total = 0
rownum = 1

for domain, prefix in DOMAIN_PREFIX.items():
    rows = controls_by_domain[domain]
    seq = 0
    for r in rows:
        seq += 1
        total += 1
        domain_counts[domain] += 1
        cid = f"{prefix}-{seq:02d}"
        (dom, sub, q, intent, ev, appl, weight, mas, rbi, sebi, cert, src) = r
        ws.append([cid, dom, sub, q, intent, ev, appl, weight, mas, rbi, sebi, cert, src])
        rownum += 1
        # framework coverage: count non-empty mapping cells
        for fw, idx in FRAMEWORKS:
            val = r[idx]
            if val and str(val).strip():
                framework_domain_counts[fw][domain] += 1
        # style row
        for ci, cell in enumerate(ws[rownum], start=1):
            cell.alignment = WRAP if ci in (4, 5, 6, 7, 9, 10, 11, 13) else TOP
            cell.border = BORDER
            cell.font = Font(size=10)
        wcell = ws.cell(row=rownum, column=8)
        wcell.alignment = CENTER
        wcell.fill = WEIGHT_FILL.get(weight, PatternFill())
        wcell.font = Font(size=10, bold=True)

# Column widths
widths = {
    "A": 12, "B": 22, "C": 24, "D": 52, "E": 40, "F": 38, "G": 30,
    "H": 11, "I": 34, "J": 36, "K": 30, "L": 18, "M": 34,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:M{rownum}"
ws.row_dimensions[1].height = 32

# ---- Sheet 2: Domains ----
ws2 = wb.create_sheet("Domains")
ws2.append(["Domain", "Prefix", "Control count", "Description"])
for c in ws2[1]:
    c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
    c.alignment = Alignment(wrap_text=True, vertical="center")
r = 1
for domain, prefix in DOMAIN_PREFIX.items():
    r += 1
    ws2.append([domain, prefix, domain_counts[domain], DOMAIN_DESC.get(domain, "")])
    for ci, cell in enumerate(ws2[r], start=1):
        cell.border = BORDER
        cell.alignment = WRAP if ci == 4 else TOP
ws2.append(["TOTAL", "", total, ""])
for cell in ws2[r + 1]:
    cell.font = Font(bold=True); cell.border = BORDER
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 10
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 80
ws2.freeze_panes = "A2"

# ---- Sheet 3: Framework Coverage ----
ws3 = wb.create_sheet("Framework Coverage")
ws3.append(["Domain"] + [fw for fw, _ in FRAMEWORKS] + ["Total controls in domain"])
for c in ws3[1]:
    c.fill = HEADER_FILL; c.font = HEADER_FONT; c.border = BORDER
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
r = 1
fw_totals = Counter()
for domain in DOMAIN_PREFIX:
    r += 1
    rowvals = [domain]
    for fw, _ in FRAMEWORKS:
        cnt = framework_domain_counts[fw][domain]
        fw_totals[fw] += cnt
        rowvals.append(cnt)
    rowvals.append(domain_counts[domain])
    ws3.append(rowvals)
    for ci, cell in enumerate(ws3[r], start=1):
        cell.border = BORDER
        cell.alignment = TOP if ci == 1 else CENTER
ws3.append(["TOTAL mapping references"] + [fw_totals[fw] for fw, _ in FRAMEWORKS] + [total])
for cell in ws3[r + 1]:
    cell.font = Font(bold=True); cell.border = BORDER
ws3.column_dimensions["A"].width = 30
for col in ("B", "C", "D", "E"):
    ws3.column_dimensions[col].width = 16
ws3.freeze_panes = "B2"
# Note row
note_r = r + 3
ws3.cell(row=note_r, column=1,
         value=("Counts show how many controls in each domain carry at least one mapping reference "
                "to that framework. A control commonly maps to multiple frameworks, so column totals "
                "exceed the unique control count."))
ws3.cell(row=note_r, column=1).alignment = WRAP
ws3.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=5)
ws3.row_dimensions[note_r].height = 45

# ---- Sheet 4: Readme ----
ws4 = wb.create_sheet("Readme")
ws4.column_dimensions["A"].width = 120
title_font = Font(bold=True, size=14, color="1F3864")
h_font = Font(bold=True, size=12, color="1F3864")
body = Font(size=11)
disc = Font(size=11, bold=True, color="C00000")

def putline(text, font=body, height=None):
    rr = ws4.max_row + 1 if ws4.max_row > 1 or ws4["A1"].value else 1
    ws4.cell(row=rr, column=1, value=text)
    ws4.cell(row=rr, column=1).font = font
    ws4.cell(row=rr, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    if height:
        ws4.row_dimensions[rr].height = height

readme_lines = [
    ("NI-TPRM Controls v2 — Third-Party Risk Control Questionnaire (Bank / FI)", title_font, 24),
    ("", body, 6),
    ("Purpose", h_font, None),
    ("A detailed, assessable third-party-risk control questionnaire for a bank / financial-institution "
     "context, written the way a bank assessor asks a vendor. Each control is mapped to the specific "
     "MAS, RBI and SEBI clause/section/notice it derives from, and tagged with inherent-risk weight, "
     "evidence expectations, applicability guidance, and recommended certification coverage.", body, 60),
    ("", body, 6),
    ("Methodology", h_font, None),
    ("1) Frameworks were re-confirmed via web research (June 2026) for current titles, dates and key "
     "requirement areas (see SOURCES.md). 2) Concrete control expectations were extracted per framework. "
     "3) Controls were authored as specific, non-duplicative vendor questions grouped by domain. "
     "4) Each control carries clause-level mappings reusing the in-app source clause IDs "
     "(web/src/data/sources/{mas,rbi,sebi}.json) so the workbook can later be reconciled with the platform.",
     body, 75),
    ("", body, 6),
    ("How this extends the current 54-control set", h_font, None),
    (f"The live platform ships a 54-row 'InfoSec Checklist' (08-June-2026 NIPL-SMICC Non-DLA Checklist) "
     f"covering storage, data privacy, technology standards, change/ops and similar themes. This v2 "
     f"workbook expands coverage to {total} controls across {len(DOMAIN_PREFIX)} domains, adding depth in "
     f"cryptography/key management, SBOM and software supply chain, sub-contracting / 4th-party, "
     f"payments-specific (PSS / PA-PG / CoFT / digital lending), cloud-specific, resilience, and "
     f"CERT-In / MAS reporting clocks — plus a regulator-agnostic baseline so 'None'/custom vendors still "
     f"receive a solid control set. It is a SUPERSET intended for review before any platform integration.",
     body, 90),
    ("", body, 6),
    ("Column guide (Controls sheet)", h_font, None),
    ("Control ID | Domain | Sub-domain | Control Question | Intent/Why | Evidence/RFI expected | "
     "Applicability guidance | Inherent-risk weight (High/Med/Low) | MAS mapping | RBI mapping | "
     "SEBI mapping | Recommended cert coverage | Source reference(s) (in-app clause IDs).", body, 45),
    ("", body, 6),
    ("Inherent-risk weight is the assessor's default importance of the control to the engagement; "
     "tune per the vendor's actual data access, criticality and applicable regulator.", body, 30),
    ("", body, 6),
    ("DISCLAIMER", disc, None),
    ("Clause and section references in this workbook were COMPILED FROM RESEARCH and the project's existing "
     "source clause library. They are indicative and MUST be verified against the official regulator PDFs "
     "(MAS / RBI / SEBI / CERT-In) before this questionnaire is used as a system of record or for any "
     "compliance attestation. Regulatory positions change: e.g., MAS Notice 655 has been cancelled/"
     "superseded (its baseline measures persist in TRM/Notice 658-era requirements); SEBI CSCRF PR.DS "
     "data-localisation provisions were reported to be kept in abeyance pending further consultation; and "
     "RBI PA Master Directions (2025) and Digital Lending Directions (2025) consolidate earlier circulars. "
     "Confirm effective dates, applicability to your entity class, and exact clause numbering before use.",
     body, 115),
    ("", body, 6),
    ("Source clause libraries reused: web/src/data/sources/mas.json, rbi.json, sebi.json. "
     "Reference list with URLs and access notes: SOURCES.md (same folder).", body, 30),
]
for i, (text, font, height) in enumerate(readme_lines):
    ws4.cell(row=i + 1, column=1, value=text)
    ws4.cell(row=i + 1, column=1).font = font
    ws4.cell(row=i + 1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    if height:
        ws4.row_dimensions[i + 1].height = height

OUT = "/home/bruhxd/Claudetest/EnhancedWorkFlow/TPRM/docs/questionnaire-v2/NI-TPRM-Controls-v2.xlsx"
wb.save(OUT)

# ---------------------------------------------------------------------------
# Verify
# ---------------------------------------------------------------------------
from openpyxl import load_workbook
v = load_workbook(OUT)
print("Saved:", OUT)
print("Sheets:", v.sheetnames)
controls_rows = v["Controls"].max_row - 1  # minus header
print("Controls rows (excl header):", controls_rows)
print("Total controls counted:", total)
print("\nPer-domain counts:")
for d in DOMAIN_PREFIX:
    print(f"  {d}: {domain_counts[d]}")
print("\nFramework mapping reference totals:", dict(fw_totals))
